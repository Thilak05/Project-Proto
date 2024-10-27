from flask import Flask, render_template, Response, jsonify, request, send_from_directory, session
import time
import pandas as pd
import cv2
import subprocess
import openpyxl
import firebase_admin
from firebase_admin import credentials, db
import json
import joblib
import os
from qr import scan_qr_code
from groq import Groq

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Set a secret key for session management

# Firebase Initialization
cred = credentials.Certificate("evolumin-fb768-firebase-adminsdk-5rhpr-15c4de4281.json")
firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://evolumin-fb768-default-rtdb.firebaseio.com/'
})

# Groq API Initialization
client = Groq(api_key='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')

EXCEL_FILE_PATH = "datas/heart_rate_data.xlsx"

def generate_frames():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()
        if not success:
            break

        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/assets/<path:filename>')
def assets(filename):
    return send_from_directory('assets', filename)

@app.route('/')
def index():
    return render_template('authenticate.html')

@app.route('/scan_qr_code', methods=['POST'])
def scan_qr_code_endpoint():
    scanned_data = scan_qr_code()
    if scanned_data:
        name = scanned_data.get("name", "N/A")
        address = scanned_data.get("address", "N/A")
        gender = "Male" if scanned_data.get("gender") == "M" else "Female"
        dob = scanned_data.get("dob", "N/A")
        
        # Set current mobile in session
        session['current_mobile'] = scanned_data.get("mobile", "N/A")  

        ref = db.reference("/users")
        ref.child(session['current_mobile']).set({
            "name": name,
            "address": address,
            "gender": gender,
            "dob": dob,
            "mobile": session['current_mobile']
        })
        return jsonify(message="Data successfully sent to Firebase")
    else:
        return jsonify(error="No QR code data found"), 400

@app.route('/start_monitoring', methods=['POST'])
def start_monitoring():
    subprocess.call(["python", "sbm.py"])
    return jsonify(status="Monitoring started")

@app.route('/get_average_bpm', methods=['GET'])
def get_average_bpm():
    current_mobile = session.get('current_mobile')  # Access mobile from session

    if not os.path.exists(EXCEL_FILE_PATH):
        return jsonify(error="BPM data not available"), 404

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    bpm_values = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=2).value]

    if not bpm_values:
        return jsonify(error="No BPM data recorded"), 404

    average_bpm = sum(bpm_values) / len(bpm_values)
    wb.close()

    if current_mobile:
        user_ref = db.reference("/users")
        user_ref.child(current_mobile).update({
            'bpm': average_bpm
        })
    
    return jsonify(average_bpm=average_bpm)

@app.route('/submit_patient_details', methods=['POST'])
def submit_patient_details():
    current_mobile = session.get('current_mobile')  # Access mobile from session
    if not current_mobile:
        return jsonify(error="No user is authenticated"), 400

    data = request.get_json()
    height = data.get('height')
    weight = data.get('weight')
    temperature = data.get('temperature')

    if not (height and weight and temperature):
        return jsonify(error="Incomplete patient details"), 400

    user_ref = db.reference("/users")
    user_ref.child(current_mobile).update({
        "height": height,
        "weight": weight,
        "temperature": temperature
    })
    
    return jsonify(message="Patient details updated successfully")

data = pd.read_excel("datas/disease_symptom_data.xlsx")
X = data.drop("Disease", axis=1)
model = joblib.load("model/disease_predictor_model.pkl")

@app.route('/symptom_questionnaire')
def symptom_questionnaire():
    return render_template('symptom_questionnaire.html')

@app.route('/question', methods=['POST'])
def question():
    responses = request.json.get('responses', {})
    remaining_symptoms = list(X.columns)

    next_symptom = next((symptom for symptom in remaining_symptoms if symptom not in responses), None)
    if not next_symptom:
        predicted_diseases = predict_disease(responses)
        return jsonify({"finished": True, "predicted_diseases": predicted_diseases})
    return jsonify({"finished": False, "next_symptom": next_symptom})

def predict_disease(responses):
    user_symptoms = [1 if responses.get(symptom) == 'Yes' else 0 for symptom in X.columns]
    predicted_disease = model.predict([user_symptoms])[0]
    prediction_proba = model.predict_proba([user_symptoms])[0]
    disease_proba_pairs = sorted(zip(model.classes_, prediction_proba), key=lambda x: x[1], reverse=True)
    top_diseases = [{"disease": disease, "probability": round(prob * 100, 2)} for disease, prob in disease_proba_pairs[:3]]
    return top_diseases

@app.route('/submit', methods=['POST'])
def submit_results():
    current_mobile = session.get('current_mobile')  # Access mobile from session
    if not current_mobile:
        return jsonify(error="No user is authenticated"), 400
    
    responses = request.json.get('responses', {})
    top_diseases = predict_disease(responses)
    
    for disease in top_diseases:
        disease["probability"] = float(disease["probability"])
    
    timestamp = int(time.time())
    
    try:
        user_ref = db.reference("/users")
        user_ref.child(current_mobile).update({
            "predicted_diseases": top_diseases,
            "timestamp": timestamp
        })
        return jsonify({"message": "Results saved successfully", "predicted_diseases": top_diseases})
    except firebase_admin.exceptions.InvalidArgumentError as e:
        return jsonify(error="Failed to save data to Firebase. Please check your data format."), 500

# New routes for user symptom input and treatment plan response
@app.route('/symptom_input')
def symptom_input():
    return render_template('symptom_input.html')

@app.route('/submit_symptoms', methods=['POST'])
def submit_symptoms():
    current_mobile = session.get('current_mobile')  # Access mobile from session
    if not current_mobile:
        return jsonify(error="No user is authenticated"), 400
    
    symptoms = request.form.get("symptoms")
    
    user_ref = db.reference(f"/users/{current_mobile}")
    user_data = user_ref.get()
    
    if user_data:
        bpm = user_data.get("bpm", "Not Available")
        temperature = user_data.get("temperature", "Not Available")
        gender = user_data.get("gender", "Not Available")
        height = user_data.get("height", "Not Available")
        weight = user_data.get("weight", "Not Available")
    else:
        return jsonify(error="Patient data not available"), 404

    # Generate prompt
    prompt = f"""
    Patient presents with the following:
    - Symptoms: {symptoms}
    - Heart Rate (BPM): {bpm}
    - Temperature: {temperature}°F
    - Gender: {gender}
    - Height: {height}
    - Weight: {weight}
    
    Based on this information, predict the disease and suggest the most appropriate medication or treatment plan.
    """

    # Query Llama model for response
    response = client.chat.completions.create(
        model="llama-3.1-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.7, 
        max_tokens=512,    
        top_p=0.9,         
        stream=False
    ).choices[0].message.content

    # Push the symptoms and prompt to Firebase
    db.reference(f"/users/{current_mobile}/symptoms").push({
        'symptoms': symptoms,
        'prompt': prompt,
    })

    return jsonify({"recommended_treatment": response})

if __name__ == '__main__':
    app.run(debug=True)
